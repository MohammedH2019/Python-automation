from email.header import Header
from operator import index
from time import time
from click import style
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils  import get_column_letter 
from typing import List
import csv
import pyexcel as pe
from copy import copy
from datetime import datetime
import datetime
import pandas as pd
# Welcome Method and ask user to input source and destination file 

csv = pd.read_csv("test1 3.csv", encoding = "utf-8")

with pd.ExcelWriter('new12.xlsx', engine = 'openpyxl') as writer:
     csv.to_excel(writer, sheet_name="Sheet1", index=False)



def welcome_Source_Destination():
    print("Welcome to the python excel automation console app")
    print("Press Enter to Continue:")
    Enter = str(input())


def source_file():
    print("Where is the source excel file (.xlxs)")
    source = str(input())
    return source

def destination_file():
    print("where is the destination excel file (.xlxs)")
    destination = str(input())

    return destination    

def destination_file_without_input():
    dest = str(destination_file())
    return dest

def option_salah():
    print("Do you want to copy Many rows and columns or just single row and column")
    response = input()
    if response == "M":
        multipleRowColumn()
    elif response == "S":
        singleRowColumn()
    else:
        print("no option chosen bye")   



# CALLING THE METHODS
def multipleRowColumn():
    selectedRange = copyMultipleRange(2,4,7,38)
    pastedRange   = pasteRange(3,123,8,153,selectedRange)

    print("all done mate")

def singleRowColumn():  
    print("Processing_fajr_salah...")
    selectedRange = copyOneRange(38,38,16,67)
    pastingRange = pasteRange(12,154,12,183,selectedRange)
    # template.save(destination_file)
    print("Range copied and pasted!")

#ORGAINSIATION 

def copyMultipleRange(startCol, startRow, endCol, endRow):
    wb = load_workbook("V2.xlsx")
    ws = wb.active 
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):

            temp_time = ws.cell(row = i, column = j).value
            print(type(temp_time))
            if isinstance(temp_time, datetime.time):
               print("variable is a time")
               time_now  = temp_time.strftime("%H:%M")
               print(time_now)
               rowSelected.append(time_now)      
            else:
              print(time_now)
              rowSelected.append(time_now)
        rangeSelected.append(rowSelected) 
    return rangeSelected




def copyOneRange(StartRow, Startcol, endcol , endRow):
    wb = load_workbook(filename = 'V2.xlsx', data_only = True )
    ws = wb.active 
    # input command to accept the arguments - int values
    rangeSelected = []
    for i in range(StartRow, endRow+1 , 1):
        print(i)
        rowSelected = []
        for j in range(endcol,endcol+1,1):
            temp_time = ws.cell(row = i, column = j).value
            print(type(temp_time))
            if isinstance(temp_time, datetime.time):
               print("variable is a time")
               time_now  = temp_time.strftime("%H:%M")
               print(time_now)
               rowSelected.append(time_now)      
            else:
              print(time_now)
              rowSelected.append(time_now)
            print(rowSelected)
        rangeSelected.append(rowSelected)
        print(rangeSelected)

    return rangeSelected   


def pasteRange(startCol, startRow, endCol, endRow,copiedData):
    template = load_workbook("new12.xlsx")
    temp_sheet = template.active
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            temp_sheet.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
    template.save("new12.xlsx")
    read_file = pd.read_excel("new12.xlsx")
    read_file.to_csv("test18.csv", index=None)



welcome_Source_Destination()
option_salah()     














